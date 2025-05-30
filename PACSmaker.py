import openpyxl
import pandas as pd
import folium
import base64
import os
from folium.plugins import Draw, LocateControl, MeasureControl
import datetime
import requests
import sys
import re
from github import Github

CURRENT_VERSION = "2.0.5"
UPDATE_DATE = "2025-05-22"
GITHUB_RAW_URL = "https://raw.githubusercontent.com/come6433/q8r2x7v1p0/main/PACSmaker.py"
REPO_NAME = 'come6433/q8r2x7v1p0'
FILENAME = "PACS.html"
IMAGES_DIR = 'images'

def get_version_from_text(text):
    m = re.search(r'CURRENT_VERSION\s*=\s*["\']([\d\.]+)["\']', text)
    return m.group(1) if m else None

def normalize_version(v):
    parts = v.split(".")
    if len(parts) == 2 and parts[1].endswith("0"):
        parts[1] = parts[1].rstrip("0")
        if parts[1] == "":
            parts[1] = "0"
    return tuple(map(int, parts))

def version_compare(a, b):
    na = normalize_version(a)
    nb = normalize_version(b)
    return (na > nb) - (na < nb)

def check_and_update():
    try:
        print("업데이트 확인 중 ...")
        r = requests.get(GITHUB_RAW_URL, timeout=5)
        if r.status_code == 200:
            remote_text = r.text
            remote_version = get_version_from_text(remote_text)
            if remote_version and version_compare(remote_version, CURRENT_VERSION) > 0:
                print(f"\n새 버전({remote_version})이 있습니다. 자동 업데이트를 진행합니다.")
                try:
                    os.rename(__file__, __file__ + ".bak")
                except Exception:
                    pass
                with open(__file__, "w", encoding="utf-8") as f:
                    f.write(remote_text)
                print("업데이트 완료! 프로그램을 다시 실행해 주세요.")
                sys.exit(0)
            else:
                print("최신 버전입니다.")
        else:
            print("업데이트 서버 연결 실패:", r.status_code)
    except Exception as e:
        print("업데이트 확인 중 오류:", e)

def print_intro():
    print("=" * 40)
    print("      PACS 저상게시대 지도 생성기")
    print("=" * 40)
    print("버전:        ", CURRENT_VERSION)
    print("업데이트:    ", UPDATE_DATE)
    print("- 민원, 예정 마커 삭제")
    print("- 설치예정, 철거예정, 변경예정 마커 추가")
    print("- 마커 라벨 길이에 따라 원 크기와 폰트 크기 자동 조정")
    print("- 범례 가운데정렬")
    print("=" * 40)

def read_excel(filename):
    print("관리목록.xlsx 파일을 읽는 중...\n")
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data = [row for row in ws.iter_rows(min_row=3, values_only=True)]
    df = pd.DataFrame(data)
    col_names = [cell.value for cell in ws[2]]
    df.columns = col_names
    df = df.dropna(subset=['설치장소', '관리번호', '위도', '경도'])
    return df

def image_to_base64(path):
    with open(path, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode()

def get_color(단수, marker_no):
    marker_no_str = str(marker_no)
    if marker_no_str.startswith('설치예정'):
        return 'yellow'  # 노란색
    if marker_no_str.startswith('철거예정'):
        return '#ff9800'  # orange
    if marker_no_str.startswith('변경예정'):
        return '#8bc34a'  # light green
    try:
        if int(단수) == 1:
            return 'blue'
        else:
            return 'red'
    except Exception:
        return 'blue'

def get_marker_text_color(bg_color):
    # 밝은 배경은 검정, 어두운 배경은 흰색
    if bg_color in ['red', 'blue', '#00bcd4', '#8bc34a', '#ff9800']:
        return 'white'
    if bg_color in ['yellow', 'pink']:
        return 'black'
    return 'black'

def make_popup_html(group, df):
    first = group.iloc[0]
    설치장소 = first['설치장소'] if '설치장소' in group.columns else ""
    popup_html = f"<div style='text-align:center;'><b class='popup-title'>{설치장소}</b><br>"
    popup_html += "<table style='border-collapse:collapse; width:auto; margin:8px auto 0 auto;'><tr>"
    for _, row in group.iterrows():
        관리번호 = str(row['관리번호'])
        image_path = f"{IMAGES_DIR}/{관리번호}.jpg"
        if os.path.exists(image_path):
            img_base64 = image_to_base64(image_path)
            popup_html += (
                f"<td style='padding:4px 8px; text-align:center;'>"
                f"<img src='data:image/jpeg;base64,{img_base64}' width='120' class='popup-img' "
                "style='cursor:zoom-in;display:block;margin:0 auto;'><br>"
                f"<span style='font-weight:bold'>{관리번호}</span></td>"
            )
        else:
            popup_html += (
                f"<td style='padding:4px 8px; text-align:center;'>"
                f"<div style='width:120px;height:90px;background:#eee;display:flex;align-items:center;justify-content:center;'>이미지 없음</div>"
                f"<br><span style='font-weight:bold'>{관리번호}</span></td>"
            )
    popup_html += "</tr></table>"
    col_count = group.shape[0]
    min_width = 120 * col_count
    popup_html += f"<table style='border-collapse:collapse; width:auto; min-width:{min_width}px; margin:8px auto 0 auto;'>"
    popup_html += "<tr><td style='border:1px solid #000; padding:4px 8px; background:#f0f0f0; font-weight:bold;'>관리번호</td>"
    for _, row in group.iterrows():
        popup_html += f"<td style='border:1px solid #000; padding:4px 8px; background:#e3f2fd; font-weight:bold;'>{row['관리번호']}</td>"
    popup_html += "</tr>"
    exclude_cols = ['마커번호', '관리번호', '위도', '경도', '설치장소', '단수', '순번']
    # Z칼럼 이후(AA~)는 무시
    max_col_index = 25  # 0부터 시작(Z=25)
    for idx, col in enumerate(df.columns):
        if col in exclude_cols:
            continue
        if idx > max_col_index:
            break
        popup_html += f"<tr><td style='border:1px solid #000; padding:4px 8px; background:#f0f0f0; font-weight:bold;'>{col}</td>"
        for _, row in group.iterrows():
            val = row[col] if pd.notnull(row[col]) else ""
            if isinstance(val, str):
                val = val.replace('\r\n', '<br>').replace('\n', '<br>')
            popup_html += f"<td style='border:1px solid #000; padding:4px 8px;'>{val}</td>"
        popup_html += "</tr>"
    popup_html += "</table><br></div>"
    return popup_html

def add_markers_to_map(m, df):
    fg1 = folium.FeatureGroup(name='1단 (파랑)').add_to(m)
    fg2 = folium.FeatureGroup(name='2단 (빨강)').add_to(m)
    fg_install = folium.FeatureGroup(name='설치예정(청록)').add_to(m)
    fg_remove = folium.FeatureGroup(name='철거예정(주황)').add_to(m)
    fg_change = folium.FeatureGroup(name='변경예정(연두)').add_to(m)

    # 카운터
    install_cnt = 1
    remove_cnt = 1
    change_cnt = 1

    grouped = df.groupby('마커번호')
    for marker_no, group in grouped:
        first = group.iloc[0]
        lat, lon = first['위도'], first['경도']
        marker_no_str = str(marker_no)
        if marker_no_str.startswith('설치예정'):
            marker_label = f"설{install_cnt}"
            install_cnt += 1
            단수 = 1
        elif marker_no_str.startswith('철거예정'):
            marker_label = f"철{remove_cnt}"
            remove_cnt += 1
            단수 = 1
        elif marker_no_str.startswith('변경예정'):
            marker_label = f"변{change_cnt}"
            change_cnt += 1
            단수 = 1
        else:
            marker_label = marker_no_str
            단수 = first['단수'] if pd.notnull(first['단수']) else 1

        # 마커 라벨 길이에 따라 원 크기와 폰트 크기 자동 조정
        label_len = len(marker_label)
        if label_len <= 2:
            size = 24
            font_size = 12
        elif label_len == 3:
            size = 28
            font_size = 12
        elif label_len == 4:
            size = 32
            font_size = 11
        else:
            size = 36
            font_size = 10

        popup_html = make_popup_html(group, df)
        bg_color = get_color(단수, marker_no)
        text_color = get_marker_text_color(bg_color)
        icon_html = (
            f"""<div style="background-color:{bg_color};color:{text_color};border-radius:50%;text-align:center;"""
            f"""width:{size}px;height:{size}px;line-height:{size}px;font-size:{font_size}px;border:1.5px solid #888;overflow:hidden;white-space:nowrap;">{marker_label}</div>"""
        )
        marker = folium.Marker(
            location=[lat, lon],
            icon=folium.DivIcon(html=icon_html),
            popup=folium.Popup(popup_html, max_width=250)
        )
        if marker_no_str.startswith('설치예정'):
            fg_install.add_child(marker)
        elif marker_no_str.startswith('철거예정'):
            fg_remove.add_child(marker)
        elif marker_no_str.startswith('변경예정'):
            fg_change.add_child(marker)
        elif int(단수) == 1:
            fg1.add_child(marker)
        else:
            fg2.add_child(marker)
    return fg1, fg2, fg_install, fg_remove, fg_change

def add_generated_time(m):
    now = datetime.datetime.now()
    time_str = f"작성시점 : {now.year}년 {now.month:02d}월 {now.day:02d}일 {now.hour:02d}시 {now.minute:02d}분"
    html = f"""<div style="position: fixed;right: 30px;bottom: 18px;background: rgba(255,255,255,0.85);color: #222;font-size: 13px;border-radius: 7px;padding: 4px 14px;box-shadow: 1px 2px 8px #bbb;z-index: 9999;pointer-events: none;">{time_str}</div>"""
    m.get_root().html.add_child(folium.Element(html))

def make_map(df):
    print("지도 작성 중 ...")
    center_lat = df.iloc[0]['위도']
    center_lon = df.iloc[0]['경도']
    m = folium.Map(location=[center_lat, center_lon], zoom_start=13, max_zoom=21, tiles=None)
    LocateControl(auto_start=False, flyTo=True, keepCurrentZoomLevel=True).add_to(m)
    MeasureControl(primary_length_unit='meters', primary_area_unit='sqmeters').add_to(m)
    vworld_base = "https://xdworld.vworld.kr/2d/Base/service/{z}/{x}/{y}.png"
    folium.TileLayer(
        tiles=vworld_base,
        attr="VWorld Base",
        name="VWorld 일반지도",
        overlay=False,
        control=True
    ).add_to(m)
    vworld_sat = "https://xdworld.vworld.kr/2d/Satellite/service/{z}/{x}/{y}.jpeg"
    folium.TileLayer(
        tiles=vworld_sat,
        attr="VWorld Satellite",
        name="VWorld 위성지도",
        overlay=False,
        control=True
    ).add_to(m)
    naver_tile = "https://map.pstatic.net/nrs/api/v1/raster/satellite/{z}/{x}/{y}.jpg?version=6.03"
    folium.TileLayer(
        tiles=naver_tile,
        attr="Naver Satellite",
        name="네이버 위성지도",
        overlay=False,
        control=True
    ).add_to(m)
    traffic_tile = "https://its.go.kr:9443/geoserver/gwc/service/wmts/rest/ntic:N_LEVEL_{z}/ntic:REALTIME/EPSG:3857/EPSG:3857:{z}/{y}/{x}?format=image/png8"
    folium.TileLayer(
        tiles=traffic_tile,
        attr="국가교통정보센터",
        name="실시간 교통정보",
        overlay=True,
        control=True,
        max_zoom=15,
        min_zoom=7,
        fmt="image/png",
        show=False
    ).add_to(m)
    fg1, fg2, fg_install, fg_remove, fg_change = add_markers_to_map(m, df)
    add_generated_time(m)
    return m

def add_legend_and_controls(m, df):
    # 범례에 신규 마커 추가, "범례"만 가운데 정렬
    df_normal = df[~df['마커번호'].astype(str).str.startswith(('설치예정', '철거예정', '변경예정'))]
    count_1 = ((df_normal['단수'] == 1)).sum()
    count_2 = ((df_normal['단수'] == 2)).sum() // 2

    count_install = df['마커번호'].astype(str).str.startswith('설치예정').sum()
    count_remove = df['마커번호'].astype(str).str.startswith('철거예정').sum()
    count_change = df['마커번호'].astype(str).str.startswith('변경예정').sum()

    legend_html = f"""
    <div id="legend" style="position: fixed; bottom: 50px; left: 50px; width: 160px; height: 140px; background-color: white; border:2px solid grey; z-index:9999; font-size:14px; padding: 10px; text-align:left;">
        <div style="text-align:center; font-weight:bold;">범례</div>
        <i style="background:blue; width:15px; height:15px; display:inline-block; border-radius:50%;"></i> 1단 - {count_1}개<br>
        <i style="background:red; width:15px; height:15px; display:inline-block; border-radius:50%;"></i> 2단 - {count_2}개<br>
        <i style="background:yellow; width:15px; height:15px; display:inline-block; border-radius:50%; border:1px solid #888;"></i> 설치예정 - {count_install}개<br>
        <i style="background:#ff9800; width:15px; height:15px; display:inline-block; border-radius:50%;"></i> 철거예정 - {count_remove}개<br>
        <i style="background:#8bc34a; width:15px; height:15px; display:inline-block; border-radius:50%;"></i> 변경예정 - {count_change}개<br>
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))
    folium.LayerControl(collapsed=False).add_to(m)

def add_custom_js_css(m):
    custom_js_css = r"""<style>
#showLatLngBtn {position: fixed;top: 20px;left: 50px;z-index: 9999;background: #1976d2;color: white;border: none;border-radius: 5px;padding: 8px 16px;font-size: 14px;cursor: pointer;box-shadow: 1px 2px 8px #888;}
#toggleBtns {position: fixed;top: 20px;left: 210px;z-index: 9999;display: flex;gap: 8px;}
#hideAllBtn, #showAllBtn {background: #e53935;color: white;border: none;border-radius: 5px;padding: 8px 16px;font-size: 14px;cursor: pointer;box-shadow: 1px 2px 8px #888;}
#showAllBtn {background: #1976d2;margin-left: 0;}
#searchBox {position: fixed;top: 70px;left: 50px;z-index: 9999;background: white;border: 1px solid #aaa;border-radius: 5px;padding: 8px 12px;width: 300px;box-shadow: 1px 2px 8px #888;}
#imgOverlay {display: none;position: fixed;z-index: 10000;left: 0; top: 0; width: 100vw; height: 100vh;background: rgba(0,0,0,0.7);justify-content: center; align-items: center;}
#imgOverlay img {max-width: 90vw; max-height: 80vh; border: 5px solid #fff; border-radius: 8px;}
#imgOverlayClose {position: absolute; top: 30px; right: 40px; color: #fff; font-size: 2em; cursor: pointer;}
</style>
<button id="showLatLngBtn" class="search-btn">위/경도 표시</button>
<span id="toggleBtns">
    <button id="hideAllBtn">모두 감추기</button>
    <button id="showAllBtn" style="display:none;">모두 보이기</button>
</span>
<div id="searchBox">
    <input id="searchInput" type="text" placeholder="설치장소, 관리부서, 관리번호 검색">
    <button id="searchBtn" class="search-btn">검색</button>
    <button id="resetBtn" class="search-btn" style="margin-left:6px;">필터초기화</button>
</div>
<div id="imgOverlay" onclick="this.style.display='none'">
    <span id="imgOverlayClose" onclick="document.getElementById('imgOverlay').style.display='none';event.stopPropagation();">&times;</span>
    <img id="imgOverlayImg" src="">
</div>
<script>
var latlngPopupActive = false;
var latlngPopup;
document.addEventListener('DOMContentLoaded', function() {
    setTimeout(function() {
        for (var key in window) {
            if (key.startsWith("map_") && window[key] instanceof L.Map) {
                window.map = window[key];
            }
        }
        if (window.map) {
            window.map.on('click', function(e) {
                if (!latlngPopupActive) return;
                if (latlngPopup) window.map.closePopup(latlngPopup);
                latlngPopup = L.popup()
                    .setLatLng(e.latlng)
                    .setContent("위도: " + e.latlng.lat.toFixed(6) + "<br>경도: " + e.latlng.lng.toFixed(6))
                    .openOn(window.map);
            });
        }
        var showLatLngBtn = document.getElementById('showLatLngBtn');
        if (showLatLngBtn) {
            showLatLngBtn.onclick = function() {
                latlngPopupActive = !latlngPopupActive;
                this.innerText = latlngPopupActive ? "위/경도 끄기" : "위/경도 표시";
                if (!latlngPopupActive && latlngPopup && window.map) {
                    window.map.closePopup(latlngPopup);
                }
            };
        }
        var allMarkers = [];
        if (window.map) {
            window.map.eachLayer(function(layer) {
                if (layer instanceof L.Marker && layer._popup) {
                    allMarkers.push(layer);
                }
            });
        }
        var searchBtn = document.getElementById('searchBtn');
        var resetBtn = document.getElementById('resetBtn');
        var searchInput = document.getElementById('searchInput');
        function filterMarkers() {
            var q = searchInput.value.trim();
            var isNumber = /^\d+$/.test(q);
            allMarkers.forEach(function(marker) {
                var html = marker._popup.getContent();
                if (typeof html !== "string" && html && html.innerHTML) {
                    html = html.innerHTML;
                }
                html = String(html).toLowerCase();
                var show = false;
                if (!q) {
                    show = true;
                } else if (isNumber) {
                    var matches = html.match(/<td[^>]*>관리번호<\/td>\s*<td[^>]*>([^<]+)<\/td>/g);
                    if (matches) {
                        for (var i = 0; i < matches.length; i++) {
                            var val = matches[i].replace(/.*<td[^>]*>관리번호<\/td>\s*<td[^>]*>([^<]+)<\/td>.*/, "$1").trim();
                            if (val === q || val.startsWith(q + "-")) {
                                show = true;
                                break;
                            }
                        }
                    }
                } else {
                    show = html.indexOf(q.toLowerCase()) !== -1;
                }
                if (show) {
                    if (!window.map.hasLayer(marker)) marker.addTo(window.map);
                } else {
                    if (window.map.hasLayer(marker)) window.map.removeLayer(marker);
                }
            });
        }
        if (searchBtn) searchBtn.onclick = filterMarkers;
        if (searchInput) searchInput.onkeydown = function(e) {
            if (e.key === "Enter") filterMarkers();
        };
        if (resetBtn) resetBtn.onclick = function() {
            searchInput.value = "";
            allMarkers.forEach(function(marker) {
                if (!window.map.hasLayer(marker)) marker.addTo(window.map);
            });
        };
    }, 300);
    var hideBtn = document.getElementById('hideAllBtn');
    var showBtn = document.getElementById('showAllBtn');
    hideBtn.onclick = function() {
        document.getElementById('showLatLngBtn').style.display = 'none';
        document.getElementById('searchBox').style.display = 'none';
        var legend = document.getElementById('legend');
        if (legend) legend.style.display = 'none';
        var layerControls = document.getElementsByClassName('leaflet-control-layers');
        for (var i = 0; i < layerControls.length; i++) {
            layerControls[i].style.display = 'none';
        }
        hideBtn.style.display = 'none';
        showBtn.style.display = '';
    };
    showBtn.onclick = function() {
        document.getElementById('showLatLngBtn').style.display = '';
        document.getElementById('searchBox').style.display = '';
        var legend = document.getElementById('legend');
        if (legend) legend.style.display = '';
        var layerControls = document.getElementsByClassName('leaflet-control-layers');
        for (var i = 0; i < layerControls.length; i++) {
            layerControls[i].style.display = '';
        }
        hideBtn.style.display = '';
        showBtn.style.display = 'none';
    };
});
document.addEventListener('click', function(e) {
    if (e.target.tagName === 'IMG' && e.target.classList.contains('popup-img')) {
        var overlay = document.getElementById('imgOverlay');
        var overlayImg = document.getElementById('imgOverlayImg');
        overlayImg.src = e.target.src;
        overlay.style.display = 'flex';
        e.stopPropagation();
    }
});
</script>
"""
    m.get_root().html.add_child(folium.Element(custom_js_css))

def save_map(m, filename):
    m.save(filename)
    print("\nHTML 파일 저장 완료:", filename)

def upload_or_update(repo, path_local, path_remote):
    print(f"업로드 시도: {path_local} -> {path_remote}")
    with open(path_local, "rb") as f:
        content = f.read()
    try:
        contents = repo.get_contents(path_remote)
        repo.update_file(path_remote, "자동 업로드", content, contents.sha)
        print(f"업데이트: {path_remote}")
    except Exception as e:
        print(f"신규 생성 시도: {path_remote} (사유: {e})")
        repo.create_file(path_remote, "자동 업로드", content)
        print(f"생성: {path_remote}")

def github_upload(filename):
    from dotenv import load_dotenv
    load_dotenv()
    GITHUB_TOKEN = os.getenv('GITHUB_TOKEN')
    SHARE_URL = f'come6433.github.io/q8r2x7v1p0/{filename}'
    g = Github(GITHUB_TOKEN)
    repo = g.get_repo(REPO_NAME)
    answer = input("\n업로드 하시겠습니까? (y/n): ").strip().lower()
    if answer == "y":
        print("\nHTML 파일 업로드 시작")
        upload_or_update(repo, filename, filename)
        excel_name = '관리목록.xlsx'
        if os.path.exists(excel_name):
            upload_or_update(repo, excel_name, excel_name)
            print(f"{excel_name} 파일도 업로드 완료!")
        else:
            print(f"{excel_name} 파일이 존재하지 않아 업로드하지 않았습니다.")
        print("\n서버 업로드 완료!")
        print(f"공유주소: {SHARE_URL}")
        print(f"※※※ 페이지가 정상적으로 표시되려면 1~2분 정도 기다려야 합니다. ※※※")
    else:
        print("\n업로드를 취소했습니다.")

def main():
    check_and_update()
    print_intro()
    df = read_excel('관리목록.xlsx')
    m = make_map(df)
    add_legend_and_controls(m, df)
    add_custom_js_css(m)
    save_map(m, FILENAME)
    github_upload(FILENAME)
    print("=" * 40)
    input("아무 키나 누르면 종료합니다.")

if __name__ == "__main__":
    main()
